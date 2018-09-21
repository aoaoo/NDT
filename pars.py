###Comments###
#Python v3.5.0
#openpyxl v2.5.3
print('***Ez a kód a PUHY-HP, -P és -EP szériák esetében számolja össze a beltéri egységekhez szükséges jointokat, szűkítőket és csöveket (azaz a kültéri egységek összekötéséhez szükséges darabokat ehhez hozzá kell venni, ha van ilyen). \nAmennyiben 56-osnál nagyobb szűkítő kellene valahol, akkor a kódot ki kell egészíteni, jelenleg nem tud annál nagyobb átmérővel dolgozni.')
print('***Jelenleg beta verzióban működik, az eredmények ellenőrzése szükséges.')
print('***A program nem a Mitsubishi szabványos elágazásait adja meg.')
print('***A végleges verzió esetén az eredmények ellenőrzése kimerülhet a Design Tool-ból exportálható és az innen számolt csőmennyiségek összehasonlításával.')
print('***Az ellenőrzésnél figyelni kell, hogy a kültériek összecsövezésére szolgáló csőmennyiség nincs benne ennek a programnak a kimenetében.')
print('***Ezek mellett a kültériek összekötésére szolgáló csövekből a NDT jellemzően az egyik géptől érkező "gáz" csövet nem számolja.')
print('***Amennyiben a G11-nél vagy a Mitsubishi New Design Tool v1.90 frissebb megjelenik, úgy ellenőrizni kell a kódba égetett adatokat.')
print('***A program a kültéri egységek összekötéséhez szükséges csőmennyiséget és jointo(ka)t nem számítja, ezt kézzel hozzá kell adni!')
print('***Használat \n1, Exportáljuk ki a NDT-ból az xml fájlt. \n2, Futtassuk ezt a programot úgy, hogy a program egy mappában van az exportált fájllal. \n3, Figyelni kell rá, hogy a programnak nincs jogosultsága létező excell fájlt felülírni, ezért [xml_filename]_joint.xlsx néven ne legyen a mappában fájl. \n4, A futás után ellenőrizzük az eredményeket: A számolt csőmennyiségeket, esetleg az egyéb kimeneti értékeit.')
print('***Forráskód: https://github.com/aoaoo/NDT')
print()
print("***Begin***")
print()

#import modules
import string #string
import re #regexp
import sys #system exit
import time #wait
from operator import itemgetter #Short lists

###Functions###

#Check extra infos (modifications because of high distances)
#The pipe diameter between the outdoor unit and the first branch
def ou_branch (si, ty): #size and type of the outdoor unit
	if ty == 'EP':
		EP_dict = {
		200: [10, 22],
		250: [10, 22],
		300: [10, 28],
		350: [12, 28],
		400: [12, 28],
		450: [15, 28],
		500: [15, 28],
		550: [15, 28],
		600: [15, 28],
		650: [15, 28],
		700: [18, 35],
		750: [18, 35],
		800: [18, 35],
		850: [18, 42],
		900: [18, 42],
		950: [18, 42],
		1000: [18, 42],
		1050: [18, 42],
		1100: [18, 42],
		1150: [18, 42],
		1200: [18, 42],
		1250: [18, 42],
		1300: [18, 42],
		1350: [18, 42]		
		}
		return EP_dict[si]
	elif ty == 'HP':
		HP_dict = {
		200: [12, 18],
		250: [12, 22],
		400: [15, 28],
		500: [15, 28]		
		}
		return HP_dict[si]	
	else: #ty == 'P'
		P_dict = {
		200: [10, 22],
		250: [10, 22],
		300: [10, 22],
		350: [12, 28],
		400: [12, 28],
		450: [15, 28],
		500: [15, 28],
		550: [15, 28],
		600: [15, 28],
		650: [15, 28],
		700: [18, 35],
		750: [18, 35],
		800: [18, 35],
		850: [18, 42],
		900: [18, 42],
		950: [18, 42],
		1000: [18, 42],
		1050: [18, 42],
		1100: [18, 42],
		1150: [18, 42],
		1200: [18, 42],
		1250: [18, 42],
		1300: [18, 42],
		1350: [18, 42]		
		}
		return P_dict[si]


#Check extra infos (modifications because of high distances)
#The pipe diameter between the indoor unit and the branch, in case of VRF indoor unit (use Miu_branch function in case of PAC-LV11)
def iu_branch (si, ty): #size of the connected indoor unit, and the type of the system (based on the data, now this second variable negigible)
	if ty == 'EP':
		EP_dict = {
		15: [6, 12],
		18: [6, 12],
		20: [6, 12],
		22: [6, 12],
		25: [6, 12],
		32: [6, 12],
		35: [6, 12],
		40: [6, 12],
		42: [6, 12],
		50: [6, 12],
		63: [10, 15],
		71: [10, 15],
		80: [10, 15],
		100: [10, 15],
		125: [10, 15],
		140: [10, 15],
		200: [10, 18],
		250: [10, 22]
		}
		return EP_dict[si]
	elif ty == 'HP':
		HP_dict = {
		15: [6, 12],
		18: [6, 12],
		20: [6, 12],
		22: [6, 12],
		25: [6, 12],
		32: [6, 12],
		35: [6, 12],
		40: [6, 12],
		42: [6, 12],
		50: [6, 12],
		63: [10, 15],
		71: [10, 15],
		80: [10, 15],
		100: [10, 15],
		125: [10, 15],
		140: [10, 15],
		200: [10, 18],
		250: [10, 22]
		}
		return HP_dict[si]
	else: #ty == 'P'
		P_dict = {
		15: [6, 12],
		18: [6, 12],
		20: [6, 12],
		22: [6, 12],
		25: [6, 12],
		32: [6, 12],
		35: [6, 12],
		40: [6, 12],
		42: [6, 12],
		50: [6, 12],
		63: [10, 15],
		71: [10, 15],
		80: [10, 15],
		100: [10, 15],
		125: [10, 15],
		140: [10, 15],
		200: [10, 18],
		250: [10, 22]
		}
		return P_dict[si]

#The pipe diameter between the indoor unit and the branch, in case of M series indoor unit (with PAC-LV11): MSZ-SF, EF, FH and MFZ-KJ
def Miu_branch (si): #In this case system type is not important
	M_dict = {
	15: [6, 10],
	18: [6, 10],
	20: [6, 10],
	22: [6, 10],
	25: [6, 10],
	30: [6, 10],
	40: [6, 10],
	45: [6, 12], #unit 50, because of the size correction
	60: [6, 15],
	71: [10, 15],
	80: [10, 15]
	}
	return M_dict[si]

#Check extra infos (modifications because of high distances)
#The pipe diameter between two branches
def branch_branch (si, ty): #size of the total down-stream indoor capacity (and the outdoor unit type)
	if ty == 'EP':
		if si <= 140:
			return [10, 15]
		elif si <= 200:
			return [10, 18]
		elif si <= 300:
			return [10, 22]
		elif si <= 400:
			return [12, 28]
		elif si <= 650:
			return [15, 28]
		elif si <= 800:
			return [18, 35]
		else:
			return [18, 42]
	elif ty == 'HP':
		if si <= 140:
			return [10, 15]
		elif si <= 200:
			return [10, 18]
		elif si <= 300:
			return [10, 22]
		elif si <= 400:
			return [12, 28]
		elif si <= 650:
			return [15, 28]
		else:
			return [-1,-1]
	else: #ty == 'P'
		if si <= 140:
			return [10, 15]
		elif si <= 200:
			return [10, 18]
		elif si <= 300:
			return [10, 22]
		elif si <= 400:
			return [12, 28]
		elif si <= 650:
			return [15, 28]
		elif si <= 800:
			return [18, 35]
		else:
			return [18, 42]

###Values from string###
def get_num (s, typ): #String and type of the expected value; need: 'import re'
	if typ == 'float':
		return float(re.search(r'[-]?\d+[.]?\d{0,3}', s.replace(',' , '.')).group(0))
	elif typ == 'int':
		return int(re.search(r'[-]?\d+', s).group(0))
	else: #positive int
		return int(re.search(r'\d+', s).group(0))

#Find if M serie (MSZ-EF/SF/FH or MFZ-KJ) (PAC-LV11 connection, to use smaller liquid pipe size)
def get_M_series (s): #String, the type of the indoor unit; need 'import re'
		if re.search(r'[M][SF][Z][-][ESFK][FHJ]', s) == None:
			return('VRF')
		else:
			return('M')

#M size correction (Because the total capacity in case of M series IU is eq or less than the number in the type designation)
def M_size_corr (size): #Size is an integer; No information for EF-18
	if size == 20 or size == 35 or size == 50:
		return size - 5
	elif size == 22 or size == 42:
		return size - 2
	else:
		return size
	
#bend length correction
def bend_len_corr (si, ty): #size of the connected indoor unit, and the type of the system
	if ty == 'EP':
		EP_dict = {
			200: 0.42,
			250: 0.42,
			300: 0.47,
			350: 0.47,
			400: 0.50,
			450: 0.50,
			500: 0.50,
			550: 0.50,
			600: 0.50,
			650: 0.50,
			700: 0.70,
			750: 0.70,
			800: 0.70,
			850: 0.80,
			900: 0.80,
			950: 0.80,
			1000: 0.80,
			1050: 0.80,
			1100: 0.80,
			1150: 0.80,
			1200: 0.80,
			1250: 0.80,
			1300: 0.80,
			1350: 0.80
		}
		return EP_dict[si]
	elif ty == 'HP':
		HP_dict = {
			200: 0.30,
			250: 0.35,
			400: 0.50,
			500: 0.50
		}
		return HP_dict[si]
	else: #ty == 'P'
		P_dict = {
			200: 0.42,
			250: 0.42,
			300: 0.47,
			350: 0.47,
			400: 0.50,
			450: 0.50,
			500: 0.50,
			550: 0.50,
			600: 0.50,
			650: 0.50,
			700: 0.70,
			750: 0.70,
			800: 0.70,
			850: 0.80,
			900: 0.80,
			950: 0.80,
			1000: 0.80,
			1050: 0.80,
			1100: 0.80,
			1150: 0.80,
			1200: 0.80,
			1250: 0.80,
			1300: 0.80,
			1350: 0.80
		}
		return P_dict[si]

#Eqvivalent distance from the first joint; Leftchild, rightchild, pipe length
def dist_1st_joint (lchild, rchild, eqlen): 
	a = len(lchild) - 1
	eqdis = [0] * (a + 1)
	for i in range (2, a + 1):
		eqdis[i] += eqlen[i]
		if lchild[i] != -1:
			eqdis[lchild[i]] += eqdis[i]
		if rchild[i] != -1:
			eqdis[rchild[i]] += eqdis[i]
	return eqdis

###Import data###
print('Írd be az importálandó .xml fájl nevét, ami ebben a mappában van: ')
while True:
	try:
		infil = str(input())
		if len(infil) > 4:
			if infil[len(infil)-4:len(infil)+1] != '.xml':
				infil = infil + '.xml'
		else:
			infil = infil + '.xml'

		import xml.etree.ElementTree as ET
		tree = ET.parse(infil)
		root = tree.getroot()
		break
	except FileNotFoundError:
		print('Írd be helyesen az importálandó .xml fájl nevét, ami ebben a mappában van: ')
	
###var
elemtyp = list() #Outdoor / Indoor, Branch
size = list() #The capacity of the units
joker1 = list() #Random joker
parentid = list()
piplen = list() #Pipe length before the unit #Expect outdoor unit
c_piplen = list() #The equivalent pipe length
numbend	= list() #Number of bent before the unit #Expect the outdoort unit
height = list() #Height of the devices
a = 0 #Where a = ID (a = NDT_ID - 1)
systyp = '' #system type: P, HP, EP
iutyp = list() #VRF or M 
dbou = 1 #Number of the outdoor unit
minh = 0 #Minimum height by the IU-s
maxh = 0 #Maximum height by the IU-s
#lenA lenB lenC nenD lenE bendA bendB bendC bendD bendE #Tho outdoor units piping length
info = '' #Informations to the excel table

for unit in root.iter('Unit'):
	elemtyp.append(unit.get('Model'))
	if elemtyp[a] == 'Outdoor':
		size.append(get_num(unit.get('ModelName'),'pint'))
		if size[0] > 700:
			info = info + ' A kültéri közeli jointokat ellenőrizni kell, hogy szükség van-e 56-osnál nagyobb szűkítőre! '
		parentid.append(-1)
		piplen.append(0)
		numbend.append(0)
		systyp = (re.search(r'[-][E,H]?[P]', unit.get('ModelName')).group(0)).replace('-', '')
		lenA = get_num(unit.get('PipeLengthA'),'float') #Because getnum change the ',' to '.' too
		bendA = get_num(unit.get('BendA'), 'pint')
		piplen.append(lenA)
		numbend.append(bendA)
		iutyp.append('VRF')
		try:
			lenB = get_num(unit.get('PipeLengthB'),'float')
			bendB = get_num(unit.get('BendB'), 'pint')
			dbou += 1
			info = info + ' Több kültéri egység esetén az őket összekötő jointokat, és csöveket a kód nem számolja, ezeket kézzel kell hozzáadni!'
			lenC = get_num(unit.get('PipeLengthC'),'float')
			bendC = get_num(unit.get('BendC'), 'pint')
			piplen[1] = lenC
			numbend[1] = bendC
			lenD = get_num(unit.get('PipeLengthD'),'float')
			bendD = get_num(unit.get('BendD'), 'pint')
			dbou += 1
			lenE = get_num(unit.get('PipeLengthE'),'float')
			bendE = get_num(unit.get('BendE'), 'pint')
			piplen[1] =  lenE
			numbend[1] = bendE
		except AttributeError:
			pass
		height.append(get_num(unit.get('Height'), 'float'))
		if dbou == 2:
			height[0] = max(height[0], get_num(unit.get('Height2'), 'float'))
		elif dbou == 3:
			height[0] = max(height[0], get_num(unit.get('Height2'), 'float'), get_num(unit.get('Height3'), 'float'))
	elif elemtyp[a] == 'Indoor':
		iutyp.append(get_M_series(unit.get('ModelName')))
		if iutyp[a] == 'M': #Size correction for PAC-LV and M series
			size.append(M_size_corr(get_num(unit.get('ModelName'),'pint')))
		else:
			size.append(get_num(unit.get('ModelName'),'pint'))
		parentid.append(int(unit.get('ParentUnitID')) - 1)
		piplen.append(get_num(unit.get('PipeLength'),'float'))
		numbend.append(get_num(unit.get('Bend'),'pint'))
		height.append(get_num(unit.get('Height'), 'float'))
		if minh == 0: #To get an existing height value for initial condition to the min and max searching method
			minh = get_num(unit.get('Height'), 'float')
			maxh = get_num(unit.get('Height'), 'float')
	elif elemtyp[a] == 'Branch': #Jump the frist branch pilength, and bend values, use lenA;;lenE and bendA;;bendE
		size.append(0)
		parentid.append(int(unit.get('ParentUnitID')) - 1)
		iutyp.append('VRF')
		if a != 1:
			piplen.append(get_num(unit.get('PipeLength'),'float'))
			numbend.append(get_num(unit.get('Bend'),'pint'))
		height.append(get_num(unit.get('Height'), 'float'))
	else:
		print('Unknown device type in sequence (0;;unit): ' + str(a))#Err mess
	bend_len_corr(200, 'EP')
	c_piplen.append(float(numbend[a]*bend_len_corr(size[0], systyp) + piplen[a]))
	a += 1
a -= 1

###Building the tree###
###var
leftsize = [0] * (a + 1) #give the capacity connected on the 1st side 
rightsize = [0] * (a + 1) #give the capacity connected on the 2nd side
leftchild = [-1] * (a + 1)  #give the just for the branch-IU connection
rightchild = [-1] * (a + 1) #just for the branch-IU connection

i = a
for i in range(a, 0, -1):
	if elemtyp[i] == 'Branch':
		size[i] = leftsize[i] + rightsize[i]
	if leftchild[parentid[i]] == -1:
		leftchild[parentid[i]] = i
		leftsize[parentid[i]] = size[i]	
	else:
		rightchild[parentid[i]] = i
		rightsize[parentid[i]] = size[i]

###Branch size calculating###
#This code calculate the amount of additional refirigent gas, and the length of the required copper pipe
#For theese feature, need to get the pipe length, and the set of the model (when two or more block are need)
#After it this code can calculate the 0th branch sizes (which connect the outdoor units if it necessary)

###var
pipdia = [[]] * (a + 1) #give diameter before of the equipment
incdia = [False] * (a + 1) #Increased the liquid pipe diameter, because height limitation or exceeding the 40m
bunit = 0 #The height of the base unit
#act_paid  actual parendtID (by function height limit test)
pipdia_len = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #6, 10 12, 15, 18, 22, 28, 35, 42, 56 #To summarize the total pipe length group by diameter
knowndia_reducer_len = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #8, 10 12, 15, 18, 22, 28, 35, 42, 56 #Pipe length to the reducer manufacturing #In this case instead of D6 use D8!!
reducer = list()
knowndia = { #known diameters to create pipdia_len vector
		6: 0,
		0: 0, #delete the zero value
		10: 1,
		12: 2,
		15: 3,
		18: 4,
		22: 5,
		28: 6,
		35: 7,
		42: 8,
		56: 9
		}

knowndia_inv = { #known diameters to create increase pipe diameter
		0: 6,
		1: 10,
		2: 12,
		3: 15,
		4: 18,
		5: 22,
		6: 28,
		7: 35,
		8: 42,
		9: 56
		}

num_join = 1 #Number of joints (int, but bool may enoug later?)

for i in range (0, a + 1):
	if elemtyp[i] == 'Outdoor':
		pipdia[i] = [0, 0]
		pass #later
	elif elemtyp[i] == 'Indoor':
		if iutyp[i] == 'M': #M serie Pipe size
			pipdia[i] = Miu_branch(size[i])
		else: #VRF pipe size
			pipdia[i] = iu_branch(size[i], systyp)
	else:
		if num_join == 1: #Not consecvent, other places the ID = 0 means OU, ID = 1 => branch
			pipdia[i] = ou_branch(size[0], systyp)
		else:
			pipdia[i] = branch_branch(size[i], systyp)
		num_join += 1

#Get the distance for every element from the first joint 
eqdist = dist_1st_joint(leftchild, rightchild, c_piplen)
#In case of P250-300, EP250-300 if the max equvivalent distance between the outdoor unit and the indoor unit is equal or more then 90/40 m, the outdoor unit liquid pipe has to reselect
if  size[0] == 250 and (systyp == 'P' or systyp == 'EP'):
	for i in eqdist:
		if i + lenA + bendA * bend_len_corr(size[0], systyp) >= 90: #Mert bugos az NDT XML exportja, és az 1. joint távolsága néha fals (lenA-val redundánsan ugyan az kéne, hogy legyen)
			pipdia[1][0] = 12
			break
if  size[0] == 300 and (systyp == 'P' or systyp == 'EP'):
	for i in eqdist:
		if i + lenA + bendA * bend_len_corr(size[0], systyp) >= 40:
			pipdia[1][0] = 12
			break

#correction of the pipe diameter (if the upper one is smaller)
for i in range(1, a + 1):
	for j in range(0, 2):
		if pipdia[leftchild[i]][j] > pipdia[i][j] and elemtyp[leftchild[i]] != 'Indoor':
			pipdia[leftchild[i]][j] = pipdia[i][j]
		elif pipdia[rightchild[i]][j] > pipdia[i][j] and elemtyp[rightchild[i]] != 'Indoor':
			pipdia[rightchild[i]][j] = pipdia[i][j]

#Correction the liquid pipe diameter, because the equvivalent distance from the first joint exceed 40m
for i in range (2, a + 1):
	if eqdist[i] > 40:
		pipdia[i][0] = knowndia_inv[knowndia[pipdia[i][0]] + 1]
		incdia[i] = True

#Correction the liquid pipe diameter, because the height difference from the mase (IU) unit is more than 15 m (and no correction was because the 40 m)
#First find the min and max value of the IU height
for i in range (2, a + 1):
	if elemtyp[i] == 'Indoor':
		if height[i] < minh:
			minh = height[i]
		elif height[i] > maxh:
			maxh = height[i]
#Choose the base IU height
if minh >= height[0]:
	bunit = minh
else:
	bunit = maxh

#Update the liquid pipe diameters
for i in range (2, a + 1):
	if elemtyp[i] == 'Indoor' and abs(height[i] - bunit) > 15 and incdia[i] == False:
		pipdia[i][0] = knowndia_inv[knowndia[pipdia[i][0]] + 1]
		incdia[i] = True
		act_paid = parentid[i]
		while abs(height[act_paid] - bunit) > 15 and act_paid != 0:
			if incdia[act_paid] == False:
				pipdia[act_paid][0] = knowndia_inv[knowndia[pipdia[act_paid][0]] + 1]
				incdia[act_paid] = True
			act_paid = parentid[act_paid]

#pipelength
for i in range(0, a + 1):
	pipdia_len[knowndia[pipdia[i][0]]] += piplen[i]
	pipdia_len[knowndia[pipdia[i][1]]] += piplen[i]

###var
b = 0 #Temp variable, joint ID (0::num_join-1)
joint = [[]] * (num_join*2) #joints
joint_s = list()

#Get the joints
joint[b] = [0,0,0]
joint[b+1] = [0,0,0]
b += 2
if size[0] >= 400:
	print('0th joint may need')#->output file
for i in range (0, a + 1):
	if elemtyp[i] == 'Branch':
		for j in range (0, 2):
			joint[b+j] = sorted([pipdia[i][j], pipdia[leftchild[i]][j], pipdia[rightchild[i]][j]], reverse=True)
		b += 2

#summarize joints
joint_s.append([joint[0][0], joint[0][1], joint[0][2], 1])
for i in range (1, len(joint)):
	exist = False
	for j in range(0, len(joint_s)):
		if joint_s[j][0:3] == joint[i]:
			joint_s[j][3] += 1
			exist = True
			break
	if exist == True:
		continue
	else:
		joint_s.append([joint[i][0], joint[i][1], joint[i][2], 1])

#calculate reducers
for i in range(1, len(joint_s)): #From 1 because the [0,0,0] joint
	for j in range(0, 2):
		knowndia_reducer_len[knowndia[joint_s[i][j]]] += 0.2*joint_s[i][3]
	knowndia_reducer_len[knowndia[joint_s[i][2]]] += 0.25*joint_s[i][3]


reducer.append([0,1,1])
for i in range (1, len(joint_s)):
	diff = 200 #lagre number
	for j in range(2, 10): #len(knowndia_inv) #8 + 8 = 16, so D8 (D6) and D10 are not interesting, but we need d12 to diff not to be 200
		if joint_s[i][1] == 6:#Use D8 instead of D6
			r1 = 8
		else:
			r1 = joint_s[i][1]
		if joint_s[i][2] == 6:
			r2 = 8
		else:
			r2 = joint_s[i][2]
		if abs(knowndia_inv[j] - (r1 + r2)) > diff or knowndia_inv[j] == 56: #What is the solution if  for example D28 + D35 has to reduce?
			passed = False
			if knowndia_inv[j] == 56 and abs(knowndia_inv[j] - (r1 + r2)) <= diff:
				for k in range(0, len(reducer)):
					if (reducer[k][0] == joint_s[i][0] or (reducer[k][0] == 8 and joint_s[i][0] == 6)) and reducer[k][1] == knowndia_inv[j]:
						reducer[k][2] += joint_s[i][3]
						passed = True
						break
			else:
				for k in range(0, len(reducer)):
					if (reducer[k][0] == joint_s[i][0] or (reducer[k][0] == 8 and joint_s[i][0] == 6)) and reducer[k][1] == knowndia_inv[j - 1]:
						reducer[k][2] += joint_s[i][3]
						passed = True
						break
			if passed == False:
				if knowndia_inv[j] == 56 and abs(knowndia_inv[j] - (r1 + r2)) <= diff:
					reducer.append([joint_s[i][0], knowndia_inv[j], joint_s[i][3]])
				else:
					reducer.append([joint_s[i][0], knowndia_inv[j - 1], joint_s[i][3]])
				break
			else:
				break #Because I cannot break from this for loop in case of True
		else:
			diff = abs(knowndia_inv[j] - (r1 + r2))


#Summarize the reducers
reducer_s = list() #Summarized reducers #delete
#reducer_pipe = [[8, 0], [10, 0], [12, 0], [15, 0], [18, 0], [22, 0], [28, 0], [35, 0], [42, 0], [56, 0]]#If the reducer is actually a straight pipe section [pipe diameter, pcs.]
reducer_s.append([0,1,1])

for i in range (1, len(reducer)):
	d1 = max(reducer[i][0], reducer[i][1])
	d2 = min(reducer[i][0], reducer[i][1])
	if d1 != d2:
		for j in range (0, len(reducer_s)):
			passed = False
			if reducer_s[j][0] == d1 and reducer_s[j][1] == d2:
				reducer_s[j][2] = reducer[i][2]
				passed = True
				break
		if passed == False:
			reducer_s.append([d1, d2, reducer[i][2]])
	else:
		print('hy')
		knowndia_reducer_len[knowndia[d1]] += 0.5

reducer_s = sorted(reducer_s, key=itemgetter(0, 1))
joint_s = sorted(joint_s, key=itemgetter(0, 1, 2))

print()
print('***End***')
print()

'''
print('___Debug___')
print(size)
print("joints: " + str(joint_s))
print('pipe length 6, 10 12, 15, 18, 22, 28, 35, 42, 56: ' + str(pipdia_len))
print('pipe length reducer 8, 10 12, 15, 18, 22, 28, 35, 42, 56: ' + str(knowndia_reducer_len))'''


#Excell manipulation
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors
infil = infil[0:len(infil)-4] + '_joint.xlsx' #Give a similar name to the output file
try: #Not use existing file (to save our files, and it's contents)
	wb_load = load_workbook(filename = infil)
	print('\n***\n A futás sikertelen.\n')
	print(infil + ' néven létezik már fájl a mappában, ennek felülírására a programnak nincs jogosultsága. \nKérlek nevezd vagy helyezd át a ' + infil +' nevű fájlt (vagy töröld), majd utána futtasd újra a kódot!\nKilépéshez a control+c gombot nyomd meg!')
	time.sleep(10)
	sys.exit()
except FileNotFoundError:
	pass

#Write an excell file
wb = Workbook()
ws = wb.active
ft_info = Font(color = colors.RED, bold = True)
ft_title = Font(bold = True)

ws['A1'] = info #Write errors, and info-s
a1 = ws['A1']
a1.font = ft_info
ws['A2'] = '1. csonk'
ws['B2'] = '2. csonk'
ws['C2'] = '3. csonk'
ws['D2'] = 'Darab'
ws.cell(column = 1, row = len(joint_s) + 5, value = 'D [mm]').font = ft_title
ws.cell(column = 2, row = len(joint_s) + 5, value = 'l_csőhöz [m]').font = ft_title
ws.cell(column = 3, row = len(joint_s) + 5, value = 'l_jointhoz [m]').font = ft_title
ws.cell(column = 4, row = len(joint_s) + 5, value = 'l_össz [m]').font = ft_title
ws.cell(column = 1, row = len(joint_s) + len(knowndia_inv) + 9, value = 'D1_szűkítő [mm]').font = ft_title
ws.cell(column = 2, row = len(joint_s) + len(knowndia_inv) + 9, value = 'D1_szűkítő [mm]').font = ft_title
ws.cell(column = 3, row = len(joint_s) + len(knowndia_inv) + 9, value = 'darab').font = ft_title
a2 = ws['A2']
b2 = ws['B2']
c2 = ws['C2']
d2 = ws['D2']
a2.font = ft_title
b2.font = ft_title
c2.font = ft_title
d2.font = ft_title
ws['A3'] = 0
ws['B3'] = 0
ws['C3'] = 0
ws['D3'] = 1
for i in range (0, len(joint_s)):
	for j in range (1, 5):
		ws.cell(column = j, row = (i + 4), value = joint_s[i][j-1])

ws.cell(column = 1, row = len(joint_s) + 6, value = knowndia_inv[0])
ws.cell(column = 2, row = len(joint_s) + 6, value = pipdia_len[0])
ws.cell(column = 3, row = len(joint_s) + 6, value = 0)
ws.cell(column = 4, row = len(joint_s) + 6, value = pipdia_len[0])

ws.cell(column = 1, row = len(joint_s) + 7, value = 8 ) #Because of the reducers
ws.cell(column = 2, row = len(joint_s) + 7, value = 0)
ws.cell(column = 3, row = len(joint_s) + 7, value = knowndia_reducer_len[0])
ws.cell(column = 4, row = len(joint_s) + 7, value = knowndia_reducer_len[0])

for i in range(1, len(knowndia_inv)):#Error, the known diameters, are not defined properly yet.
	ws.cell(column = 1, row = len(joint_s) + 7 + i, value = knowndia_inv[i])
	ws.cell(column = 2, row = len(joint_s) + 7 + i, value = pipdia_len[i])
	ws.cell(column = 3, row = len(joint_s) + 7 + i, value = knowndia_reducer_len[i])
	ws.cell(column = 4, row = len(joint_s) + 7 + i, value = pipdia_len[i] + knowndia_reducer_len[i])

for i in range(0, len(reducer)):#Error, the known diameters, are not defined properly yet.
	ws.cell(column = 1, row = len(joint_s) + len(knowndia_inv) + 10 + i, value = reducer_s[i][0])
	ws.cell(column = 2, row = len(joint_s) + len(knowndia_inv) + 10 + i, value = reducer_s[i][1])
	ws.cell(column = 3, row = len(joint_s) + len(knowndia_inv) + 10 + i, value = reducer_s[i][2])
	
wb.save(infil)

print('A program futása sikeres volt. \nKilépéshez nyomd meg a control+c gombot, vagy várj 5 másodpercet!')
time.sleep(5)
sys.exit()


'''
print('****Test section starts here******')
print('****End******')
'''
#bent -> pipe length in the export?
#outdoor unit joint calculate #Bug: In case of two (or more?) OU one of them pipe doesent calculate properly (just the liquid pipe summarize in the excell export, but the gas pipe is not summarized)
#-1 return value -> error message; + do not cause error
#Reducer diameters: 6, 8, 10, 12, 16, 18 (soft);22, 28, 35, 42, 56
#Reducer pipe: 2x20cm + from the smallest is 25cm
#If reducer D1 == reducer D2, than use ??m pipe